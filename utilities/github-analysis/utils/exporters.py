#!/usr/bin/env python3
"""
Data export utilities for CSV, Markdown, JSON, and Excel formats
"""

import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    openpyxl = None
    Font = None
    PatternFill = None
    Alignment = None
    dataframe_to_rows = None
    pd = None
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)

def save_to_csv(data: List[Dict[str, Any]], output_file: str, fieldnames: Optional[List[str]] = None) -> None:
    """Save data to CSV file with error handling"""
    if not data:
        logger.warning("No data to save")
        print("No data to save")
        return

    if not isinstance(data, list):
        raise ValueError("Data must be a list")

    try:
        # Create directory if it doesn't exist
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        # Auto-detect fieldnames if not provided
        if not fieldnames and isinstance(data[0], dict):
            fieldnames = ['index'] + list(data[0].keys())

        with open(output_file, 'w', newline='', encoding='utf-8') as file:
            if fieldnames:
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()

                for i, item in enumerate(data, 1):
                    try:
                        # Add index if not present
                        if 'index' not in item:
                            row = {'index': i, **item}
                        else:
                            row = item
                        writer.writerow(row)
                    except Exception as e:
                        logger.error(f"Error writing row {i}: {e}")
                        continue
            else:
                # Fallback for non-dict data
                writer = csv.writer(file)
                writer.writerows(data)

        print(f"✓ Saved {len(data)} items to {output_file}")
        logger.info(f"Successfully saved {len(data)} items to {output_file}")

    except Exception as e:
        error_msg = f"Failed to save CSV file {output_file}: {e}"
        logger.error(error_msg)
        raise Exception(error_msg)

def save_to_markdown(data, output_file, title, metadata=None, columns=None):
    """Save data to Markdown file with table format"""
    if not data:
        print("No data to save")
        return

    with open(output_file, 'w', encoding='utf-8') as file:
        # Title
        file.write(f"# {title}\n\n")

        # Metadata section
        if metadata:
            for key, value in metadata.items():
                file.write(f"- **{key}**: {value}\n")
            file.write("\n---\n\n")

        # Table
        if isinstance(data[0], dict):
            # Use provided columns or auto-detect
            if not columns:
                columns = [k for k in data[0].keys() if k != 'index']

            # Table header
            headers = ['#'] + [col.replace('_', ' ').title() for col in columns]
            file.write('| ' + ' | '.join(headers) + ' |\n')
            file.write('|' + '|'.join(['---'] * len(headers)) + '|\n')

            # Table rows
            for i, item in enumerate(data, 1):
                row_data = [str(i)]
                for col in columns:
                    value = item.get(col, 'N/A')
                    # Handle special formatting for links
                    if col == 'title' and 'url' in item:
                        value = f"[{value}]({item['url']})"
                    elif col == 'issue_number':
                        value = f"#{value}"
                    row_data.append(str(value))

                file.write('| ' + ' | '.join(row_data) + ' |\n')

    print(f"✓ Saved {len(data)} items to {output_file}")

def save_to_json(data: List[Dict[str, Any]], output_file: str) -> None:
    """Save data to JSON format with error handling"""
    if not data:
        logger.warning("No data to save")
        print("No data to save")
        return

    try:
        # Create directory if it doesn't exist
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)

        print(f"✓ Saved {len(data)} items to {output_file}")
        logger.info(f"Successfully saved {len(data)} items to {output_file}")

    except Exception as e:
        error_msg = f"Failed to save JSON file {output_file}: {e}"
        logger.error(error_msg)
        print(f"❌ Error saving JSON: {e}")
        raise Exception(error_msg)

def save_to_excel(data: List[Dict[str, Any]], output_file: str, sheet_name: str = 'Analysis') -> None:
    """Save data to Excel format with formatting and error handling"""
    if not EXCEL_AVAILABLE:
        error_msg = "Excel export requires: pip install openpyxl pandas"
        logger.error(error_msg)
        print(f"❌ {error_msg}")
        raise ImportError(error_msg)

    if not data:
        logger.warning("No data to save")
        print("No data to save")
        return

    try:
        # Create directory if it doesn't exist
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        # Convert to DataFrame
        df = pd.DataFrame(data)

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add data to worksheet
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[2]:  # Row 2 contains headers
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = int(adjusted_width)

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        stats = generate_summary_stats(data)

        summary_ws.append(["Metric", "Value"])
        for key, value in stats.items():
            summary_ws.append([key, value])

        # Format summary sheet
        for cell in summary_ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        wb.save(output_file)
        print(f"✓ Saved {len(data)} items to {output_file} with summary sheet")
        logger.info(f"Successfully saved {len(data)} items to {output_file}")

    except Exception as e:
        error_msg = f"Failed to save Excel file {output_file}: {e}"
        logger.error(error_msg)
        print(f"❌ Error saving Excel: {e}")
        raise Exception(error_msg)

def generate_summary_stats(data):
    """Generate summary statistics for data"""
    if not data:
        return {}

    stats = {
        'Total Items': len(data),
        'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # Add specific stats based on data type
    if isinstance(data[0], dict):
        # Count by state if available
        if 'state' in data[0]:
            states = {}
            for item in data:
                state = item.get('state', 'unknown')
                states[state] = states.get(state, 0) + 1

            for state, count in states.items():
                stats[f"{state.title()} Items"] = count

        # Average age if available
        if 'age_days' in data[0]:
            ages = [item.get('age_days', 0) for item in data if isinstance(item.get('age_days'), (int, float))]
            if ages:
                stats['Average Age (days)'] = int(round(sum(ages) / len(ages), 1))

    return stats
